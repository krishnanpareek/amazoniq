"""Extract confirmed MRTS NSA series from the Census year-sheet workbook.

Does not interpolate. Suppressed cells stay as the literal source token (e.g. '(S)').
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from category_config import CATEGORIES  # noqa: E402

RAW_XLSX = ROOT / "data" / "raw" / "census" / "mrtssales92-present.xlsx"
OUT_CSV = ROOT / "data" / "processed" / "fact_census_retail_nsa.csv"
TARGET_NAICS = {c["census_naics"] for c in CATEGORIES}
NAICS_TO_CATEGORY = {c["census_naics"]: c for c in CATEGORIES}


def _month_headers(header_row: tuple) -> list[tuple[int, str]]:
    months: list[tuple[int, str]] = []
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        text = str(value).strip()
        if text.startswith(("Jan.", "Feb.", "Mar.", "Apr.", "May", "Jun.", "Jul.", "Aug.", "Sep.", "Oct.", "Nov.", "Dec.")):
            months.append((idx, text))
    return months


def _parse_month_label(label: str, sheet_year: str) -> tuple[str, bool]:
    preliminary = "(p)" in label.lower()
    clean = label.replace("(p)", "").replace("(r)", "").strip().rstrip(".")
    month_token = clean.split()[0].replace(".", "")
    month_map = {
        "Jan": "01",
        "Feb": "02",
        "Mar": "03",
        "Apr": "04",
        "May": "05",
        "Jun": "06",
        "Jul": "07",
        "Aug": "08",
        "Sep": "09",
        "Oct": "10",
        "Nov": "11",
        "Dec": "12",
    }
    month = month_map[month_token]
    return f"{sheet_year}-{month}-01", preliminary


def extract() -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)
    rows_out: list[dict[str, object]] = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        ws = wb[sheet_name]
        header_row = None
        in_not_adjusted = False
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            kind = str(values[1]).strip() if values[1] is not None else ""
            if any(v is not None and str(v).startswith("Jan.") for v in values):
                header_row = values
            if kind.upper() == "NOT ADJUSTED":
                in_not_adjusted = True
                continue
            if kind.upper().startswith("ADJUSTED") and "NOT" not in kind.upper():
                in_not_adjusted = False
                continue
            if not in_not_adjusted or header_row is None:
                continue
            naics = str(values[0]).strip() if values[0] is not None else ""
            if naics not in TARGET_NAICS:
                continue
            months = _month_headers(tuple(header_row))
            meta = NAICS_TO_CATEGORY[naics]
            for col_idx, label in months:
                period, preliminary = _parse_month_label(label, sheet_name)
                raw = values[col_idx]
                if raw is None or raw == "":
                    sales = None
                    sales_note = "missing"
                elif isinstance(raw, str):
                    sales = None
                    sales_note = raw.strip()
                else:
                    sales = int(raw) if float(raw) == int(raw) else float(raw)
                    sales_note = "reported"
                rows_out.append(
                    {
                        "period": period,
                        "category_id": meta["category_id"],
                        "category_name": meta["name"],
                        "census_naics": naics,
                        "census_label": meta["census_label"],
                        "kind_of_business_source": kind,
                        "sales_millions_nsa": sales,
                        "sales_note": sales_note,
                        "preliminary_flag": preliminary,
                        "source_id": "SRC-CENSUS-001",
                        "source_file": RAW_XLSX.name,
                        "source_sheet": sheet_name,
                    }
                )
    rows_out.sort(key=lambda r: (r["category_id"], r["period"]))
    return rows_out


def write_csv(rows: list[dict[str, object]]) -> None:
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    extracted = extract()
    write_csv(extracted)
    print(f"wrote {len(extracted)} rows -> {OUT_CSV}")
    by_cat = {}
    for row in extracted:
        by_cat.setdefault(row["category_id"], 0)
        by_cat[row["category_id"]] += 1
    print(by_cat)
