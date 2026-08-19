"""One-off inspector for the Census MRTS workbook structure."""

from pathlib import Path

import openpyxl

path = Path(__file__).resolve().parents[1] / "data" / "raw" / "census" / "mrtssales92-present.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames)
for name in ("2026", "2025"):
    ws = wb[name]
    print(f"\n=== {name} ===")
    for i, row in enumerate(ws.iter_rows(max_row=40, max_col=16, values_only=True), start=1):
        cells = ["" if c is None else str(c)[:60] for c in row]
        print(f"{i:02d} | " + " || ".join(cells))
